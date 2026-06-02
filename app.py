# main.py
import ast
import json
from typing import Any, Dict

import uvicorn
from fastapi import FastAPI
from hdbcli import dbapi
from pydantic import BaseModel
import os
from dotenv import load_dotenv
from datetime import datetime
from Prompts.prompts import merge_prompts
from finance_guardrails import check_query, process_query
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

from logger import get_logger
logger = get_logger("percibot")

def main():
    logger.info("Application started")
    logger.warning("This is a warning")
    logger.error("Something went wrong")

# =========================
# Excel logging config + helper
# =========================
REPORT_XLSX = os.getenv("PERCIBOT_REPORT_XLSX", "perci_query_log.xlsx")
SHEET_NAME = os.getenv("PERCIBOT_REPORT_SHEET", "Query_Log")

REPORT_COLUMNS = [
    "timestamp",
    "user_query",
    "guardrails_result",           # Query is valid / invalid / error string
    "llm_sql_raw",                # raw L1 output
    "l1_time_sec",                # inference time for SQL generation
    "l1_error",                   # errors during SQL generation/parsing/execution
    "llm_answer",                 # L2 answer text
    "l2_time_sec",                # inference time for answer generation
    "l2_error",
    "sql_tokens"
    "answer_tokens"# errors during L2 generation
]
# stored_queries if needed

# Load environment variables
load_dotenv()

connection = dbapi.connect(
    address=os.getenv('HANA_HOST'),
    port=int(os.getenv('HANA_PORT')),
    user=os.getenv('HANA_USER'),
    password=os.getenv('HANA_PASSWORD'),
    encrypt=True,
    sslValidateCertificate = True,
    sslCryptoProvider = "openssl"
)
cursor = connection.cursor()
open_sql_schema = "SPNI_REPORTING#SPNIREPORTING"
space_schema= "SPNI_REPORTING"

api_key = os.getenv("OPENAI_API_KEY")
if api_key:
    print(f"API Key loaded: {api_key[:10]}...")
else:
    print("OPENAI_API_KEY not found. Continuing without LLM...")

from openai import OpenAI

client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
MODEL = os.getenv("OPENAI_MODEL", "gpt-4o-mini")


def generate_sql_with_llm(merged_prompt: str, user_question: str):
    resp = client.responses.create(
        model=MODEL,
        input=merged_prompt + "\n\nUser question:\n" + user_question +
              "\n\nReturn ONLY the SAP HANA SQL query in a single code block. No explanation.",
    )
    # # created_at= datetime(resp.created_at)
    # # completed_at = int(resp.completed_at)
    # # print(created_at,completed_at)
    #
    # # LLM-side inference time (seconds)
    # inference_sec = None
    # if isinstance(created_at, int) and isinstance(completed_at, int):
    #     inference_sec = completed_at - created_at
    # print(inference_sec)
    print(resp)
    created = datetime.fromtimestamp(resp.created_at)

    completed = datetime.fromtimestamp(resp.completed_at)

    inference_sec= resp.completed_at - resp.created_at
    sql_tokens =resp.usage.total_tokens
    return resp.output[0].content[0].text,inference_sec,sql_tokens#(resp.output_text or "").strip()



def generate_answers_with_llm(merged_prompt: str, user_question: str, df:str):
    resp = client.responses.create(
        model=MODEL,
        input=merged_prompt + "\n\nUser question:\n" + user_question + df +
              "\n\nReturn well-formed summarized answers.",
    )
    # created_at = getattr(resp, "created_at", None)
    # completed_at = getattr(resp, "completed_at", None)
    #
    # # LLM-side inference time (seconds)
    # inference_sec = None
    # if isinstance(created_at, int) and isinstance(completed_at, int):
    #     inference_sec = completed_at - created_at

    created = datetime.fromtimestamp(resp.created_at)

    completed = datetime.fromtimestamp(resp.completed_at)

    inference_sec= resp.completed_at - resp.created_at
    answer_tokens = resp.usage.total_tokens

    return resp.output[0].content[0].text,inference_sec,answer_tokens#(resp.output_text or "").strip()



def append_row_to_excel(row: Dict[str, Any]) -> None:
    """
    Appends ONE row to a single Excel file (creates file+headers if missing).
    Never creates a new file per request — always appends.
    """
    if not os.path.exists(REPORT_XLSX):
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(REPORT_COLUMNS)

        wrap = Alignment(wrap_text=True, vertical="top")
        for cell in ws[1]:
            cell.alignment = wrap

        wb.save(REPORT_XLSX)

    wb = load_workbook(REPORT_XLSX)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.create_sheet(SHEET_NAME)

    # ensure header row exists
    if ws.max_row == 0:
        ws.append(REPORT_COLUMNS)

    ws.append([row.get(col, "") for col in REPORT_COLUMNS])

    # wrap the newly added row for readability
    wrap = Alignment(wrap_text=True, vertical="top")
    for cell in ws[ws.max_row]:
        cell.alignment = wrap

    wb.save(REPORT_XLSX)


app = FastAPI()


class QueryInput(BaseModel):
    query: str


@app.post("/ask_finance")
def ask_finance(body: QueryInput):
    row: Dict[str, Any] = {k: "" for k in REPORT_COLUMNS}
    row["timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    row["user_query"] = body.query

    llm_sql_raw = "none"
    llm_answer = ""
    answer = ""
    l1_err = ""
    l2_err = ""
    sql_tokens=""
    answer_tokens=""
    inference_sec1=""
    inference_sec2=""
    try:
        print(f"Received: {body.query}")

        answer = process_query(body.query)
        print(answer)
        row["guardrails_result"] = answer
        llm_sql = "none"
        llm_answer= ""
        df=""
        if answer == "Query is valid":
            merged_prompt = merge_prompts(stage="sql")
            print(merged_prompt)  # <-- removed comma, now it's a string

            try:
                llm_sql, inference_sec1,sql_tokens = generate_sql_with_llm(merged_prompt, body.query)
            except Exception as e:
                print("Error generating sql query", e)
                l1_err = e
            row["llm_sql_raw"] = llm_sql
            row["l1_time_sec"] = inference_sec1
            row["l1_error"] = l1_err
            print(type(llm_sql),llm_sql)
            result = ast.literal_eval(str(llm_sql))
            # result[1] = result[1].replace("'","")
            print(type(result),result)
            if result[0]=="SQL QUERY":
                cursor.execute(result[1])
                sql_result = cursor.fetchall()
                df=str(sql_result)
                merged_prompt = merge_prompts(stage="answer",user_query=body.query, df=df, sql_query=result[1])
                try:
                    llm_answers,inference_sec2, answer_tokens = generate_answers_with_llm(merged_prompt, body.query, df)
                except Exception as e:
                    print("Error generating answers", e)
                    l2_err = f"L1_GENERATION_ERROR: {e}"
                print(type(llm_answers),llm_answers)
                row["llm_answer"] = llm_answers
                row["l2_time_sec"] = inference_sec2
                row["l2_error"] = l1_err
                row["sql_token"] =sql_tokens
                row["answer_tokens"]= answer_tokens
            else:
                print(result)
                print(result)
                llm_answer=result[1]
                row["llm_sql_raw"] = llm_sql
                row["l1_time_sec"] = inference_sec1
                row["l1_error"] = l1_err
                row["sql_token"] = sql_tokens
                row["answer_tokens"] = answer_tokens
            # cursor.close()
        # connection.close()
        append_row_to_excel(row)
        return {

            "query": body.query,
            "answer": llm_sql,
            "llm_answer": llm_answers,
            "status": "success",
            "validation": "valid"
        }

    except Exception as e:
        # connection.close()
        print(f"BLOCKED: {e}")
        row["guardrails_result"] = answer or "error"
        row["l1_error"] = row.get("l1_error", "")
        row["l2_error"] = row.get("l2_error", "")
        row["l1_error"] = (row["l1_error"] + f" | ENDPOINT_ERROR: {e}").strip()

        append_row_to_excel(row)

        return {
            "query": body.query,
            "answer": answer,
            "llm_answer": llm_sql,
            "status": "blocked",
            "validation": "invalid",
            "reason": str(e)
        }


# @app.get("/queries")
# def get_queries():
#     """Return all stored queries (in-memory)."""
#     return {
#         "queries": stored_queries,
#         "total": len(stored_queries)
#     }


@app.get("/")
def root():
    """Health check"""
    return {"message": "Finance Guardrails API running!"}




